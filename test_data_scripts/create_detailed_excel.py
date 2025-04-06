import pandas as pd
import datetime
import random
import os
import hashlib
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a directory for data if it doesn't exist
if not os.path.exists('data'):
    os.makedirs('data')

# Function to generate a 10-digit checksum from product details
def generate_checksum(mfg_date, exp_date, factory_location, manufacturer, item_name, mrp, more_details):
    combined_data = f"{mfg_date}&{exp_date}&{factory_location}&{manufacturer}&{item_name}&{mrp}&{more_details}"
    hash_object = hashlib.md5(combined_data.encode())
    hash_hex = hash_object.hexdigest()
    checksum = int(hash_hex[:10], 16) % 10000000000
    return f"{checksum:010d}"

# Function to generate a 10-digit unique number
def generate_unique_number():
    return f"{random.randint(1, 9999999999):010d}"

# Function to format Excel worksheet
def format_worksheet(worksheet, df):
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Format headers
    for idx, col in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = alignment
        
        # Adjust column width
        max_length = max(
            df[col].astype(str).apply(len).max(),
            len(col)
        )
        worksheet.column_dimensions[get_column_letter(idx)].width = max_length + 2
    
    # Format data cells
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = alignment

# Sample data lists
factory_locations = [
    'HP2 Plant', 'Mumbai Plant', 'Bangalore Plant', 'Delhi Plant', 'Chennai Plant',
    'Hyderabad Plant', 'Kolkata Plant', 'Pune Plant', 'Ahmedabad Plant', 'Lucknow Plant'
]

manufacturers = [
    'Ranbaxy', 'Cipla', 'Sun Pharma', 'Dr. Reddy\'s', 'Aurobindo Pharma',
    'Lupin', 'Glenmark', 'Cadila Healthcare', 'Biocon', 'Divis Laboratories'
]

item_names = [
    'Crocin 500mg', 'Paracetamol 650mg', 'Amoxicillin 250mg', 'Omeprazole 20mg', 'Metformin 500mg',
    'Aspirin 75mg', 'Cetirizine 10mg', 'Pantoprazole 40mg', 'Azithromycin 500mg', 'Metoprolol 25mg'
]

more_details = [
    'Temperature sensitive', 'Keep in cool place', 'Store in dry place', 'Avoid direct sunlight',
    'Keep away from children', 'For external use only', 'Take with food', 'Take after meals',
    'Do not exceed recommended dose', 'Prescription required'
]

# First, create the product database
product_data = []
for _ in range(10):
    mfg_date = (datetime.datetime.now() - datetime.timedelta(days=random.randint(30, 365))).strftime('%Y-%m-%d')
    exp_date = (datetime.datetime.now() + datetime.timedelta(days=random.randint(365, 730))).strftime('%Y-%m-%d')
    factory_location = random.choice(factory_locations)
    manufacturer = random.choice(manufacturers)
    item_name = random.choice(item_names)
    mrp = round(random.uniform(10, 500), 2)
    details = random.choice(more_details)
    
    # Generate source info (checksum)
    source_info = generate_checksum(mfg_date, exp_date, factory_location, manufacturer, item_name, mrp, details)
    
    # Generate object code and unique number for each type (Carton, Box, Item)
    carton_code = f"{source_info}C{generate_unique_number()}"
    box_code = f"{source_info}B{generate_unique_number()}"
    item_code = f"{source_info}I{generate_unique_number()}"
    
    product_data.append({
        'QR_Code_Carton': carton_code,
        'QR_Code_Box': box_code,
        'QR_Code_Item': item_code,
        'Date_of_MFG': mfg_date,
        'Date_of_EXP': exp_date,
        'Origin_Factory_Location': factory_location,
        'Name_of_manufacturer': manufacturer,
        'Name_of_item': item_name,
        'MRP': mrp,
        'More_Details': details,
        'SourceInfo': source_info
    })

# Create product database DataFrame
product_df = pd.DataFrame(product_data)

# Save product database to Excel with a new filename
product_excel_path = 'data/product_database_new.xlsx'

try:
    with pd.ExcelWriter(product_excel_path, engine='openpyxl') as writer:
        product_df.to_excel(writer, index=False, sheet_name='Product Database')
        worksheet = writer.sheets['Product Database']
        format_worksheet(worksheet, product_df)
    
    print(f"Product database Excel file created successfully at {product_excel_path}")
    print("\nSample entries:")
    for i, row in product_df.iterrows():
        print(f"\nEntry {i+1}:")
        print(f"Product: {row['Name_of_item']} by {row['Name_of_manufacturer']}")
        print(f"SourceInfo: {row['SourceInfo']}")
        print(f"Carton QR: {row['QR_Code_Carton']}")
        print(f"Box QR: {row['QR_Code_Box']}")
        print(f"Item QR: {row['QR_Code_Item']}")
        
except Exception as e:
    print(f"Error creating product database Excel file: {e}")

# Now create the detailed inventory using the same product data
inventory_data = []
for product in product_data:
    # Generate random quantities
    num_cartons = random.randint(5, 20)
    num_boxes = random.randint(20, 50)
    num_items = random.randint(100, 500)
    
    # Add row for Carton QR Code
    inventory_data.append({
        'SourceInfo': product['SourceInfo'],
        'QR_Code': product['QR_Code_Carton'],
        'QR_Code_Type': 'Carton',
        'Date_of_MFG': product['Date_of_MFG'],
        'Date_of_EXP': product['Date_of_EXP'],
        'Origin_Factory_Location': product['Origin_Factory_Location'],
        'Name_of_Manufacturer': product['Name_of_manufacturer'],
        'Name_of_Item': product['Name_of_item'],
        'MRP': product['MRP'],
        'More_Details': product['More_Details'],
        'Quantity': num_cartons
    })
    
    # Add row for Box QR Code
    inventory_data.append({
        'SourceInfo': product['SourceInfo'],
        'QR_Code': product['QR_Code_Box'],
        'QR_Code_Type': 'Box',
        'Date_of_MFG': product['Date_of_MFG'],
        'Date_of_EXP': product['Date_of_EXP'],
        'Origin_Factory_Location': product['Origin_Factory_Location'],
        'Name_of_Manufacturer': product['Name_of_manufacturer'],
        'Name_of_Item': product['Name_of_item'],
        'MRP': product['MRP'],
        'More_Details': product['More_Details'],
        'Quantity': num_boxes
    })
    
    # Add row for Item QR Code
    inventory_data.append({
        'SourceInfo': product['SourceInfo'],
        'QR_Code': product['QR_Code_Item'],
        'QR_Code_Type': 'Item',
        'Date_of_MFG': product['Date_of_MFG'],
        'Date_of_EXP': product['Date_of_EXP'],
        'Origin_Factory_Location': product['Origin_Factory_Location'],
        'Name_of_Manufacturer': product['Name_of_manufacturer'],
        'Name_of_Item': product['Name_of_item'],
        'MRP': product['MRP'],
        'More_Details': product['More_Details'],
        'Quantity': num_items
    })

# Create inventory DataFrame
inventory_df = pd.DataFrame(inventory_data)

# Save inventory details to Excel with a new filename
inventory_excel_path = 'data/product_inventory_details_new.xlsx'

try:
    with pd.ExcelWriter(inventory_excel_path, engine='openpyxl') as writer:
        inventory_df.to_excel(writer, index=False, sheet_name='Inventory Details')
        worksheet = writer.sheets['Inventory Details']
        format_worksheet(worksheet, inventory_df)
    
    print(f"\nInventory details Excel file created successfully at {inventory_excel_path}")
    print("\nSample entries:")
    for i, row in inventory_df.iterrows():
        if i % 3 == 0:  # Print header for each product
            print(f"\nProduct {(i//3)+1}:")
            print(f"SourceInfo: {row['SourceInfo']}")
            print(f"Product: {row['Name_of_Item']} by {row['Name_of_Manufacturer']}")
        print(f"{row['QR_Code_Type']} QR Code: {row['QR_Code']} (Quantity: {row['Quantity']})")
        
except Exception as e:
    print(f"Error creating inventory details Excel file: {e}") 