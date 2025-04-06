import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a directory for data if it doesn't exist
if not os.path.exists('data'):
    os.makedirs('data')

# Function to format Excel worksheet
def format_worksheet(worksheet, df):
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Format headers
    for idx, col in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = alignment
        
        # Adjust column width
        max_length = max(
            df[col].astype(str).apply(len).max(),
            len(col)
        )
        worksheet.column_dimensions[get_column_letter(idx)].width = max_length + 2
    
    # Format data cells
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = alignment

# Function to create an empty validation tracker
def create_empty_tracker():
    tracker_path = 'data/qr_validation_tracker.xlsx'
    
    # Create empty DataFrame with all required columns
    validation_df = pd.DataFrame(columns=[
        'QR_Code',
        'Source_Info',
        'Distributor_Reached_Status',
        'Distributor_Geo_Location',
        'Distributor_Time_Stamp',
        'Dealer_Reached_Status',
        'Dealer_Geo_Location',
        'Dealer_Time_Stamp',
        'Retailer_Status',
        'Retailer_Geo_Location',
        'Retailer_Time_Stamp'
    ])
    
    # Save empty tracker
    with pd.ExcelWriter(tracker_path, engine='openpyxl') as writer:
        validation_df.to_excel(writer, index=False, sheet_name='QR Validation Tracker')
        worksheet = writer.sheets['QR Validation Tracker']
        format_worksheet(worksheet, validation_df)
    
    print(f"Empty QR validation tracker created at {tracker_path}")
    return validation_df

# Function to populate tracker with data from inventory
def populate_tracker_from_inventory():
    tracker_path = 'data/qr_validation_tracker.xlsx'
    
    # Load inventory database
    try:
        inventory_df = pd.read_excel('data/product_inventory_details_new.xlsx')
    except FileNotFoundError:
        print("Error: Inventory file not found at data/product_inventory_details_new.xlsx")
        return None
    
    # Create entries for each QR code in the inventory
    entries = []
    for _, row in inventory_df.iterrows():
        entry = {
            'QR_Code': row['QR_Code'],
            'Source_Info': row['SourceInfo'],
            'Distributor_Reached_Status': 'Not Reached',
            'Distributor_Geo_Location': '',
            'Distributor_Time_Stamp': '',
            'Dealer_Reached_Status': 'Not Reached',
            'Dealer_Geo_Location': '',
            'Dealer_Time_Stamp': '',
            'Retailer_Status': 'Not Reached',
            'Retailer_Geo_Location': '',
            'Retailer_Time_Stamp': ''
        }
        entries.append(entry)
    
    # Create DataFrame from entries
    validation_df = pd.DataFrame(entries)
    
    # Save populated tracker
    with pd.ExcelWriter(tracker_path, engine='openpyxl') as writer:
        validation_df.to_excel(writer, index=False, sheet_name='QR Validation Tracker')
        worksheet = writer.sheets['QR Validation Tracker']
        format_worksheet(worksheet, validation_df)
    
    print(f"QR validation tracker populated with {len(entries)} entries from inventory")
    return validation_df

# Example usage:
if __name__ == "__main__":
    # Create empty tracker
    create_empty_tracker()
    
    # Populate tracker with data from inventory
    populated_df = populate_tracker_from_inventory()
    
    if populated_df is not None:
        print("\nSample entries from populated tracker:")
        for i, row in populated_df.head(3).iterrows():
            print(f"\nEntry {i+1}:")
            print(f"QR Code: {row['QR_Code']}")
            print(f"Source Info: {row['Source_Info']}")
            print(f"Distributor Status: {row['Distributor_Reached_Status']}")
            print(f"Dealer Status: {row['Dealer_Reached_Status']}")
            print(f"Retailer Status: {row['Retailer_Status']}")
        
        print(f"\nTotal entries in tracker: {len(populated_df)}") 